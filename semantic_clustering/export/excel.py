"""
Excel export functionality for semantic keyword clustering.
"""

import os
import logging
from typing import Dict, List, Optional, Tuple, Any, Union
import tempfile
from datetime import datetime
import pandas as pd
import numpy as np

# Check if openpyxl is installed
try:
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.drawing.image import Image as XLImage
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

logger = logging.getLogger(__name__)

def check_dependencies() -> bool:
    """
    Check if all required dependencies for Excel export are installed.
    
    Returns:
        True if all dependencies are installed, False otherwise
    """
    if not OPENPYXL_AVAILABLE:
        logger.error("openpyxl is required for Excel export. Install it using 'pip install openpyxl'")
        return False
    
    return True

def create_cluster_worksheet(
    workbook: 'openpyxl.Workbook',
    clusters: Dict[str, List[str]],
    cluster_labels: Optional[Dict[str, str]] = None,
    sheet_name: str = "Clusters"
) -> Optional['openpyxl.worksheet.worksheet.Worksheet']:
    """
    Create a worksheet with clustering results.
    
    Args:
        workbook: Openpyxl workbook to add the worksheet to
        clusters: Dictionary of cluster_id -> list of keywords
        cluster_labels: Dictionary of cluster_id -> descriptive label (optional)
        sheet_name: Name of the worksheet
        
    Returns:
        Created worksheet, or None if worksheet creation failed
    """
    if not check_dependencies():
        return None
    
    try:
        # Create worksheet
        if sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
        else:
            sheet = workbook.create_sheet(title=sheet_name)
        
        # Define styles
        header_font = Font(bold=True, size=12, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        cluster_id_font = Font(bold=True, size=11)
        cluster_label_font = Font(italic=True, size=11)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Set column widths
        sheet.column_dimensions['A'].width = 15
        sheet.column_dimensions['B'].width = 30
        sheet.column_dimensions['C'].width = 50
        
        # Add header
        headers = ["Cluster ID", "Cluster Label", "Keywords"]
        for col, header in enumerate(headers, 1):
            cell = sheet.cell(row=1, column=col)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border
        
        # Sort clusters by size (descending)
        sorted_clusters = sorted(
            clusters.items(),
            key=lambda x: len(x[1]),
            reverse=True
        )
        
        # Add data
        row = 2
        for cluster_id, keywords in sorted_clusters:
            # Get cluster label
            cluster_label = cluster_labels.get(cluster_id, "") if cluster_labels else ""
            
            # Add cluster ID
            cell_id = sheet.cell(row=row, column=1)
            cell_id.value = cluster_id
            cell_id.font = cluster_id_font
            cell_id.alignment = Alignment(horizontal='center', vertical='center')
            cell_id.border = border
            
            # Add cluster label
            cell_label = sheet.cell(row=row, column=2)
            cell_label.value = cluster_label
            cell_label.font = cluster_label_font
            cell_label.alignment = Alignment(horizontal='left', vertical='center')
            cell_label.border = border
            
            # Add keywords (sorted and joined with commas)
            cell_keywords = sheet.cell(row=row, column=3)
            cell_keywords.value = ", ".join(sorted(keywords))
            cell_keywords.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
            cell_keywords.border = border
            
            # Increase row height based on number of keywords
            sheet.row_dimensions[row].height = max(20, min(15 * (len(keywords) / 5), 150))
            
            row += 1
        
        # Freeze the header row
        sheet.freeze_panes = "A2"
        
        return sheet
        
    except Exception as e:
        logger.error(f"Error creating cluster worksheet: {e}")
        return None

def create_metrics_worksheet(
    workbook: 'openpyxl.Workbook',
    evaluation_metrics: Dict[str, Any],
    sheet_name: str = "Evaluation Metrics"
) -> Optional['openpyxl.worksheet.worksheet.Worksheet']:
    """
    Create a worksheet with evaluation metrics.
    
    Args:
        workbook: Openpyxl workbook to add the worksheet to
        evaluation_metrics: Dictionary of evaluation metrics
        sheet_name: Name of the worksheet
        
    Returns:
        Created worksheet, or None if worksheet creation failed
    """
    if not check_dependencies() or not evaluation_metrics:
        return None
    
    try:
        # Create worksheet
        if sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
        else:
            sheet = workbook.create_sheet(title=sheet_name)
        
        # Define styles
        header_font = Font(bold=True, size=12, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        metric_font = Font(size=11)
        value_font = Font(size=11)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Set column widths
        sheet.column_dimensions['A'].width = 30
        sheet.column_dimensions['B'].width = 20
        
        # Add header
        headers = ["Metric", "Value"]
        for col, header in enumerate(headers, 1):
            cell = sheet.cell(row=1, column=col)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border
        
        # Filter and sort metrics
        sorted_metrics = []
        for key, value in evaluation_metrics.items():
            # Skip complex data structures
            if isinstance(value, (dict, list, tuple, set, np.ndarray)):
                continue
            
            # Format the key for display
            display_key = key.replace("_", " ").title()
            
            # Format the value for display
            if isinstance(value, float):
                display_value = f"{value:.4f}"
            else:
                display_value = str(value)
            
            sorted_metrics.append((display_key, display_value))
        
        # Sort by metric name
        sorted_metrics.sort(key=lambda x: x[0])
        
        # Add data
        for row, (metric, value) in enumerate(sorted_metrics, 2):
            # Add metric name
            cell_metric = sheet.cell(row=row, column=1)
            cell_metric.value = metric
            cell_metric.font = metric_font
            cell_metric.alignment = Alignment(horizontal='left', vertical='center')
            cell_metric.border = border
            
            # Add metric value
            cell_value = sheet.cell(row=row, column=2)
            cell_value.value = value
            cell_value.font = value_font
            cell_value.alignment = Alignment(horizontal='right', vertical='center')
            cell_value.border = border
        
        # Freeze the header row
        sheet.freeze_panes = "A2"
        
        return sheet
        
    except Exception as e:
        logger.error(f"Error creating metrics worksheet: {e}")
        return None

def create_summary_worksheet(
    workbook: 'openpyxl.Workbook',
    clusters: Dict[str, List[str]],
    evaluation_metrics: Optional[Dict[str, Any]] = None,
    sheet_name: str = "Summary"
) -> Optional['openpyxl.worksheet.worksheet.Worksheet']:
    """
    Create a summary worksheet.
    
    Args:
        workbook: Openpyxl workbook to add the worksheet to
        clusters: Dictionary of cluster_id -> list of keywords
        evaluation_metrics: Dictionary of evaluation metrics (optional)
        sheet_name: Name of the worksheet
        
    Returns:
        Created worksheet, or None if worksheet creation failed
    """
    if not check_dependencies():
        return None
    
    try:
        # Create worksheet
        if sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
        else:
            sheet = workbook.create_sheet(title=sheet_name)
        
        # Define styles
        header_font = Font(bold=True, size=12)
        value_font = Font(size=11)
        title_font = Font(bold=True, size=14)
        
        # Set column widths
        sheet.column_dimensions['A'].width = 30
        sheet.column_dimensions['B'].width = 20
        
        # Add title
        cell = sheet.cell(row=1, column=1)
        cell.value = "Semantic Keyword Clustering Summary"
        cell.font = title_font
        sheet.merge_cells('A1:B1')
        cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Add date
        cell = sheet.cell(row=2, column=1)
        cell.value = "Generated on:"
        cell.font = header_font
        
        cell = sheet.cell(row=2, column=2)
        cell.value = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        cell.font = value_font
        
        # Add cluster statistics
        row = 4
        
        # Add header
        cell = sheet.cell(row=row, column=1)
        cell.value = "Cluster Statistics"
        cell.font = header_font
        row += 1
        
        # Number of clusters
        cell = sheet.cell(row=row, column=1)
        cell.value = "Number of Clusters"
        cell.font = value_font
        
        cell = sheet.cell(row=row, column=2)
        cell.value = len(clusters)
        cell.font = value_font
        row += 1
        
        # Total keywords
        total_keywords = sum(len(keywords) for keywords in clusters.values())
        
        cell = sheet.cell(row=row, column=1)
        cell.value = "Total Keywords"
        cell.font = value_font
        
        cell = sheet.cell(row=row, column=2)
        cell.value = total_keywords
        cell.font = value_font
        row += 1
        
        # Average keywords per cluster
        cell = sheet.cell(row=row, column=1)
        cell.value = "Average Keywords per Cluster"
        cell.font = value_font
        
        cell = sheet.cell(row=row, column=2)
        cell.value = f"{total_keywords / len(clusters):.2f}" if clusters else "0"
        cell.font = value_font
        row += 1
        
        # Min and max cluster sizes
        if clusters:
            cluster_sizes = [len(keywords) for keywords in clusters.values()]
            min_size = min(cluster_sizes)
            max_size = max(cluster_sizes)
            
            cell = sheet.cell(row=row, column=1)
            cell.value = "Smallest Cluster Size"
            cell.font = value_font
            
            cell = sheet.cell(row=row, column=2)
            cell.value = min_size
            cell.font = value_font
            row += 1
            
            cell = sheet.cell(row=row, column=1)
            cell.value = "Largest Cluster Size"
            cell.font = value_font
            
            cell = sheet.cell(row=row, column=2)
            cell.value = max_size
            cell.font = value_font
            row += 1
        
        # Add evaluation metrics if provided
        if evaluation_metrics:
            row += 1
            
            cell = sheet.cell(row=row, column=1)
            cell.value = "Evaluation Metrics"
            cell.font = header_font
            row += 1
            
            # Add key metrics
            key_metrics = [
                "silhouette_score",
                "calinski_harabasz_score",
                "davies_bouldin_score"
            ]
            
            for metric in key_metrics:
                if metric in evaluation_metrics:
                    value = evaluation_metrics[metric]
                    if isinstance(value, float):
                        value = f"{value:.4f}"
                    
                    cell = sheet.cell(row=row, column=1)
                    cell.value = metric.replace("_", " ").title()
                    cell.font = value_font
                    
                    cell = sheet.cell(row=row, column=2)
                    cell.value = value
                    cell.font = value_font
                    row += 1
        
        # Make this the first sheet
        workbook.active = sheet
        
        return sheet
        
    except Exception as e:
        logger.error(f"Error creating summary worksheet: {e}")
        return None

def export_to_excel(
    clusters: Dict[str, List[str]],
    output_path: str,
    cluster_labels: Optional[Dict[str, str]] = None,
    evaluation_metrics: Optional[Dict[str, Any]] = None,
    include_summary: bool = True,
    include_metrics: bool = True
) -> bool:
    """
    Export clustering results to Excel.
    
    Args:
        clusters: Dictionary of cluster_id -> list of keywords
        output_path: Path to save the Excel file
        cluster_labels: Dictionary of cluster_id -> descriptive label (optional)
        evaluation_metrics: Dictionary of evaluation metrics (optional)
        include_summary: Whether to include a summary worksheet
        include_metrics: Whether to include a metrics worksheet
        
    Returns:
        True if Excel export was successful, False otherwise
    """
    if not check_dependencies():
        return False
    
    try:
        # Create workbook
        workbook = openpyxl.Workbook()
        
        # Remove default sheet
        default_sheet = workbook.active
        if "Sheet" in workbook.sheetnames:
            workbook.remove(default_sheet)
        
        # Create worksheets
        if include_summary:
            create_summary_worksheet(workbook, clusters, evaluation_metrics)
        
        create_cluster_worksheet(workbook, clusters, cluster_labels)
        
        if include_metrics and evaluation_metrics:
            create_metrics_worksheet(workbook, evaluation_metrics)
        
        # Save workbook
        workbook.save(output_path)
        
        return True
        
    except Exception as e:
        logger.error(f"Error exporting to Excel: {e}")
        return False

def clusters_to_dataframe(
    clusters: Dict[str, List[str]],
    cluster_labels: Optional[Dict[str, str]] = None
) -> pd.DataFrame:
    """
    Convert clusters to a pandas DataFrame.
    
    Args:
        clusters: Dictionary of cluster_id -> list of keywords
        cluster_labels: Dictionary of cluster_id -> descriptive label (optional)
        
    Returns:
        DataFrame with cluster information
    """
    rows = []
    
    for cluster_id, keywords in clusters.items():
        label = cluster_labels.get(cluster_id, "") if cluster_labels else ""
        
        for keyword in keywords:
            rows.append({
                "cluster_id": cluster_id,
                "cluster_label": label,
                "keyword": keyword
            })
    
    return pd.DataFrame(rows)
